"""
Exchange rates, sourced from the National Bank of Romania.

Two things need rates, and they need them differently:

  * Logging an expense in a currency that isn't the account's default. This
    has to work on every request, offline-ish, and give the same answer twice
    - so it reads a snapshot from disk (see RATES_DIR) rather than calling BNR
    inline. A weekly automation refreshes that folder.

    Two file shapes are read from it, newest first: the .xlsx the scheduled
    automation drops there (Monedă / Curs (RON) / Dată, one row per currency)
    and the .json this module writes itself. The spreadsheet is the one a
    person can open and read, so it wins when it is the more recent of the
    two; the JSON carries all 37 BNR currencies rather than the five on the
    sheet, so it stays useful as a fallback.
  * The rates page, which just shows today's numbers.

BNR publishes one reference rate per currency per banking day, as XML, at
curs.bnr.ro. Rates are quoted as "how many RON for one unit", except for a
handful of small-value currencies quoted per 100 or per 10000 (the
`multiplier` attribute) - those are normalised to per-unit here so nothing
downstream has to know the difference.

BNR does NOT publish cryptocurrency rates - it's a central bank, and its list
is 37 fiat currencies plus gold (XAU) and IMF drawing rights (XDR). BTC and
ETH therefore come from a public crypto price API, and their RON leg is still
BNR's: price-in-USD from the crypto source, USD-to-RON from BNR. The snapshot
records which source each rate came from so the UI can say so rather than
implying the central bank blesses a bitcoin price.
"""

import json
import os
import re
import urllib.request
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree

BNR_XML_URL = "https://curs.bnr.ro/nbrfxrates.xml"
CRYPTO_URL = (
    "https://api.coingecko.com/api/v3/simple/price"
    "?ids=bitcoin,ethereum&vs_currencies=usd&include_last_updated_at=true"
)

BASE_CURRENCY = "RON"

# The currencies the app offers when logging an expense, and the ones the
# rates page shows. Everything BNR publishes is stored in the snapshot
# regardless - these are just what the UI puts in front of you.
SUPPORTED_CURRENCIES = ["RON", "EUR", "USD", "GBP"]
CRYPTO_CURRENCIES = {"BTC": "bitcoin", "ETH": "ethereum"}
PAGE_CURRENCIES = ["EUR", "USD", "GBP", "BTC", "ETH"]

# Where the weekly automation drops its snapshots. Overridable so tests (and
# a container that mounts the folder somewhere else) don't write into the repo.
RATES_DIR = Path(os.environ.get("BNR_RATES_DIR") or (Path(__file__).resolve().parent.parent / "cursuri_bnr"))

LATEST_FILE = "latest.json"

# How out of date the on-disk snapshot may be before a read falls back to
# calling BNR directly. The automation runs weekly, BNR publishes on banking
# days only, and a long holiday can stretch a gap - so this allows a missed
# run plus a quiet week rather than tripping on the first late Monday.
MAX_SNAPSHOT_AGE_DAYS = 16

HTTP_TIMEOUT_SECONDS = 15

_XML_NS = {"bnr": "https://www.bnr.ro/xsd"}

# A live fetch is only ever a fallback, but a cold or stale deploy would
# otherwise call BNR on every single request. One process-wide cache entry is
# enough: the rate changes once a day.
_live_cache: Optional[dict] = None
_live_cache_at: Optional[datetime] = None
LIVE_CACHE_SECONDS = 3600


class RatesUnavailable(Exception):
    """No snapshot on disk and BNR couldn't be reached."""


# --- fetching --------------------------------------------------------------


def _http_get(url: str) -> bytes:
    # BNR serves the XML to anything, but a default urllib User-Agent is the
    # kind of thing that gets rate-limited first, so it identifies itself.
    request = urllib.request.Request(url, headers={"User-Agent": "expense-tracker/1.0 (+bnr rates)"})
    with urllib.request.urlopen(request, timeout=HTTP_TIMEOUT_SECONDS) as response:
        return response.read()


def parse_bnr_xml(xml_bytes: bytes) -> dict:
    """
    Turns BNR's daily XML into {"date": "YYYY-MM-DD", "rates": {code: ron_per_unit}}.

    The published number is RON per `multiplier` units (multiplier defaults to
    1), so it's divided back down to one unit here - otherwise HUF, quoted per
    100, would come out a hundred times too expensive.
    """
    root = ElementTree.fromstring(xml_bytes)

    cube = root.find(".//bnr:Body/bnr:Cube", _XML_NS)
    if cube is None:
        raise ValueError("BNR XML has no Cube element - the feed format changed")

    rates: dict[str, float] = {BASE_CURRENCY: 1.0}
    for rate in cube.findall("bnr:Rate", _XML_NS):
        code = (rate.get("currency") or "").upper()
        if not code or not rate.text:
            continue
        multiplier = float(rate.get("multiplier") or 1)
        try:
            value = float(rate.text)
        except ValueError:
            continue
        if multiplier <= 0:
            continue
        rates[code] = value / multiplier

    if "EUR" not in rates:
        raise ValueError("BNR XML parsed but carries no EUR rate - refusing a half-empty snapshot")

    return {"date": cube.get("date") or date.today().isoformat(), "rates": rates}


def fetch_crypto(ron_per_usd: float) -> dict:
    """
    BTC/ETH priced in RON. The crypto source quotes USD; the USD->RON leg is
    BNR's own rate, so the RON figure is only ever one hop from the reference
    rate the rest of the app uses.

    Returns {} rather than raising: a crypto API being down is not a reason to
    lose the fiat rates the expense flow actually depends on.
    """
    try:
        payload = json.loads(_http_get(CRYPTO_URL))
    except Exception:
        return {}

    out: dict[str, dict] = {}
    for code, coin_id in CRYPTO_CURRENCIES.items():
        entry = payload.get(coin_id) or {}
        usd = entry.get("usd")
        if not usd:
            continue
        out[code] = {
            "ron": float(usd) * ron_per_usd,
            "usd": float(usd),
            "updated_at": entry.get("last_updated_at"),
        }
    return out


def build_snapshot(*, include_crypto: bool = True) -> dict:
    """Fetches everything the app shows, in one pass, and shapes it for disk."""
    fiat = parse_bnr_xml(_http_get(BNR_XML_URL))
    crypto = fetch_crypto(fiat["rates"].get("USD", 0.0)) if include_crypto else {}

    return {
        "date": fiat["date"],  # the banking day BNR published for, not the day we fetched
        "fetched_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "base": BASE_CURRENCY,
        "source": BNR_XML_URL,
        "crypto_source": CRYPTO_URL.split("?")[0] if crypto else None,
        "rates": fiat["rates"],
        "crypto": crypto,
    }


# --- the snapshot on disk --------------------------------------------------


def save_snapshot(snapshot: dict, directory: Optional[Path] = None) -> Path:
    """
    Writes the snapshot twice: once under its banking date (so the folder is a
    history you can look back through) and once as latest.json (so readers
    don't have to sort filenames).
    """
    target = Path(directory) if directory else RATES_DIR
    target.mkdir(parents=True, exist_ok=True)

    body = json.dumps(snapshot, indent=2, ensure_ascii=False, sort_keys=True) + "\n"
    dated = target / f"{snapshot['date']}.json"
    dated.write_text(body, encoding="utf-8")
    (target / LATEST_FILE).write_text(body, encoding="utf-8")
    return dated


# Dates as they appear in the folder's filenames: the ISO ones this module
# writes ("2026-09-04.json") and the Romanian ones the scheduled automation
# writes ("Curs BNR 06.09.2026.xlsx").
_FILENAME_DATES = (
    (re.compile(r"(\d{4})-(\d{2})-(\d{2})"), (1, 2, 3)),
    (re.compile(r"(\d{2})[.\-](\d{2})[.\-](\d{4})"), (3, 2, 1)),
)


def _date_from_name(path: Path) -> Optional[date]:
    for pattern, (y, m, d) in _FILENAME_DATES:
        found = pattern.search(path.name)
        if not found:
            continue
        try:
            return date(int(found.group(y)), int(found.group(m)), int(found.group(d)))
        except ValueError:
            continue
    return None


def _parse_sheet_date(value) -> Optional[str]:
    """The Dată column, which arrives as either a real date or 'DD.MM.YYYY' text."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if not value:
        return None
    for pattern, (y, m, d) in _FILENAME_DATES:
        found = pattern.search(str(value))
        if not found:
            continue
        try:
            return date(int(found.group(y)), int(found.group(m)), int(found.group(d))).isoformat()
        except ValueError:
            continue
    return None


def read_xlsx_snapshot(path: Path) -> Optional[dict]:
    """
    Reads the spreadsheet the scheduled automation writes: a header row, then
    one row per currency of (code, rate in RON, date).

    Deliberately loose about layout - it takes the first cell of a row as the
    code and the first numeric cell after it as the rate, so a reordered or
    re-titled column doesn't silently produce an empty snapshot. A row whose
    code isn't a currency (the header, a blank spacer, a total) simply doesn't
    parse and is skipped.

    Crypto rows are separated out from fiat ones the same way a fetched
    snapshot separates them, so everything downstream - conversion, the rates
    page's source labels - can't tell which format it came from.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    try:
        sheet = workbook.active
        rates: dict[str, float] = {BASE_CURRENCY: 1.0}
        crypto: dict[str, dict] = {}
        published: Optional[str] = None

        for row in sheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            code = str(row[0]).strip().upper()
            if not (2 <= len(code) <= 5) or not code.isalpha():
                continue

            value = next(
                (float(cell) for cell in row[1:]
                 if isinstance(cell, (int, float)) and not isinstance(cell, bool) and float(cell) > 0),
                None,
            )
            if value is None:
                continue

            if code in CRYPTO_CURRENCIES:
                crypto[code] = {"ron": value}
            else:
                rates[code] = value

            published = published or next((_parse_sheet_date(cell) for cell in row[1:] if _parse_sheet_date(cell)), None)
    finally:
        workbook.close()

    if len(rates) <= 1 and not crypto:
        return None

    return {
        # The sheet's own date if it carries one, else the day in its filename -
        # a snapshot with no date would read as infinitely stale.
        "date": published or (_date_from_name(path) or date.today()).isoformat(),
        "base": BASE_CURRENCY,
        "source": str(path.name),
        "rates": rates,
        "crypto": crypto,
    }


def _read_any(path: Path) -> Optional[dict]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return read_xlsx_snapshot(path)
    try:
        snapshot = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return snapshot if isinstance(snapshot, dict) and snapshot.get("rates") else None


def load_snapshot(directory: Optional[Path] = None) -> Optional[dict]:
    """
    The newest snapshot in the folder, whichever format it is in, or None if
    there is nothing readable there.

    Ordering is by the date in the filename rather than by name, because the
    two naming conventions in the folder don't sort against each other -
    "Curs BNR 06.09.2026.xlsx" would otherwise lose to "2026-09-04.json".
    Files with no date in the name fall back to their modification time, and
    latest.json - which by design never has one - is read for its own date.
    """
    target = Path(directory) if directory else RATES_DIR
    if not target.is_dir():
        return None

    candidates: list[tuple[date, Path]] = []
    for path in target.iterdir():
        if not path.is_file() or path.suffix.lower() not in (".json", ".xlsx", ".xlsm"):
            continue
        if path.name.startswith("~$"):
            continue  # Excel's lock file for a sheet someone has open
        stamp = _date_from_name(path)
        if stamp is None and path.name == LATEST_FILE:
            written = _read_any(path)
            stamp = _date_from_name(Path(str((written or {}).get("date", "")))) if written else None
        if stamp is None:
            stamp = date.fromtimestamp(path.stat().st_mtime)
        candidates.append((stamp, path))

    # Newest first; on the same day prefer the spreadsheet, since that is the
    # one the scheduled automation just wrote.
    candidates.sort(key=lambda pair: (pair[0], pair[1].suffix.lower() in (".xlsx", ".xlsm")), reverse=True)

    for _, path in candidates:
        snapshot = _read_any(path)
        if snapshot:
            return snapshot
    return None


def snapshot_age_days(snapshot: dict) -> Optional[int]:
    try:
        published = date.fromisoformat(snapshot["date"])
    except (KeyError, TypeError, ValueError):
        return None
    return (date.today() - published).days


def is_stale(snapshot: dict) -> bool:
    age = snapshot_age_days(snapshot)
    return age is None or age > MAX_SNAPSHOT_AGE_DAYS


# --- what the rest of the app calls ----------------------------------------


def _live_snapshot() -> Optional[dict]:
    global _live_cache, _live_cache_at

    now = datetime.now(timezone.utc)
    if _live_cache and _live_cache_at and (now - _live_cache_at) < timedelta(seconds=LIVE_CACHE_SECONDS):
        return _live_cache

    try:
        snapshot = build_snapshot()
    except Exception:
        return None

    snapshot["live"] = True
    _live_cache, _live_cache_at = snapshot, now

    # Best-effort write-through: on a host where the folder is writable this
    # doubles as a cache across restarts. On a read-only or ephemeral disk it
    # simply doesn't happen, which costs nothing.
    try:
        save_snapshot(snapshot)
    except OSError:
        pass

    return snapshot


def get_snapshot(*, allow_live: bool = True) -> dict:
    """
    The rates to use right now: the committed snapshot when it's current,
    otherwise a live fetch, otherwise the stale snapshot with a flag on it.

    Serving stale-but-real rates beats failing outright - a three-week-old EUR
    rate converts an expense to within a percent or so, and the response says
    how old it is so the UI can tell you.
    """
    snapshot = load_snapshot()

    if snapshot and not is_stale(snapshot):
        return {**snapshot, "stale": False}

    if allow_live:
        live = _live_snapshot()
        if live:
            return {**live, "stale": False}

    if snapshot:
        return {**snapshot, "stale": True, "age_days": snapshot_age_days(snapshot)}

    raise RatesUnavailable("No exchange rates on disk and BNR could not be reached")


def rate_to_ron(currency: str, snapshot: dict) -> float:
    """How many RON one unit of `currency` is worth."""
    code = (currency or "").strip().upper()
    if code == BASE_CURRENCY:
        return 1.0

    crypto = (snapshot.get("crypto") or {}).get(code)
    if crypto and crypto.get("ron"):
        return float(crypto["ron"])

    rate = (snapshot.get("rates") or {}).get(code)
    if not rate:
        raise KeyError(code)
    return float(rate)


def convert(amount: float, source: str, target: str, snapshot: Optional[dict] = None) -> tuple[float, float]:
    """
    Converts between any two currencies in the snapshot, via RON.

    Returns (converted_amount, rate) where rate is target units per one source
    unit - stored alongside the expense so a past conversion stays explainable
    after the rate has moved on.
    """
    source = (source or "").strip().upper()
    target = (target or "").strip().upper()
    if source == target:
        return round(amount, 2), 1.0

    snapshot = snapshot or get_snapshot()
    rate = rate_to_ron(source, snapshot) / rate_to_ron(target, snapshot)
    return round(amount * rate, 2), rate


def page_rates(snapshot: Optional[dict] = None) -> dict:
    """
    The rates page's payload: EUR/USD/GBP/BTC/ETH against RON, each labelled
    with where it actually came from.
    """
    snapshot = snapshot or get_snapshot()
    rows = []

    for code in PAGE_CURRENCIES:
        if code in CRYPTO_CURRENCIES:
            entry = (snapshot.get("crypto") or {}).get(code)
            if not entry:
                continue
            row = {"currency": code, "ron": round(float(entry["ron"]), 2), "source": "crypto", "kind": "crypto"}
            # The spreadsheet only carries a RON figure; a fetched snapshot also
            # knows the USD one. Omit rather than send null, so the frontend can
            # test for the key.
            if entry.get("usd"):
                row["usd"] = entry["usd"]
            rows.append(row)
        else:
            value = (snapshot.get("rates") or {}).get(code)
            if not value:
                continue
            rows.append({
                "currency": code,
                "ron": round(float(value), 4),
                "source": "bnr",
                "kind": "fiat",
            })

    return {
        "date": snapshot.get("date"),
        "fetched_at": snapshot.get("fetched_at"),
        "base": BASE_CURRENCY,
        "stale": bool(snapshot.get("stale")),
        "age_days": snapshot.get("age_days"),
        "rates": rows,
    }
