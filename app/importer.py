"""
Parses an uploaded budget spreadsheet (the same shape as the user's own
"<Month> Budget Tracker" Google Sheets) into structured rows the API can
upsert into the database.

The sheet's layout is two parallel tables sharing a row range:

    INCOME name, INCOME amount | SPENDINGS name, amount, tag ("Needs"/"Wants")

Once one side runs out of rows the other side keeps going, and a handful of
label rows (TOTAL INCOME, TOTAL EXPENSES, Ramas, GOALS/ACTUAL blocks) are
interleaved rather than confined to fixed columns. Rather than hard-coding
column positions (fragile across exports), this scans each row's non-empty
cells left to right and classifies (name, amount[, tag]) groups by shape:
a tag-less pair is income, a tag'd pair/triple is a spending category, and
the GOALS/ACTUAL block is state-tracked across rows since its WANTS/NEEDS/
SAVINGS rows don't repeat the header.

Noise rows (stray text with no parseable amount) are skipped, not errored -
these spreadsheets are hand-edited and occasionally have a stray keystroke
row.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl import load_workbook

from app.utils import parse_european_amount

GOAL_TAGS = ("WANTS", "NEEDS", "SAVINGS")
CATEGORY_TAGS = ("Needs", "Wants", "Savings")

MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
FILENAME_PERIOD = re.compile(r"(20\d{2}).*?\b(" + "|".join(MONTH_ABBR) + r")[a-z]*\b", re.IGNORECASE)


@dataclass
class IncomeRow:
    name: str
    amount: float


@dataclass
class CategoryRow:
    name: str
    amount: float
    tag: Optional[str]  # "Needs" / "Wants" / "Savings"


@dataclass
class GoalPct:
    target_pct: Optional[float] = None
    actual_pct: Optional[float] = None


@dataclass
class ImportResult:
    income: List[IncomeRow] = field(default_factory=list)
    categories: List[CategoryRow] = field(default_factory=list)
    goals: dict = field(default_factory=dict)  # {"Wants": GoalPct, "Needs": GoalPct, "Savings": GoalPct}
    warnings: List[str] = field(default_factory=list)


def guess_period_from_filename(filename: str) -> Optional[str]:
    """"2026 Aug Budget.xlsx" -> "2026-08"; None if the name doesn't match."""
    if not filename:
        return None
    match = FILENAME_PERIOD.search(filename)
    if not match:
        return None
    year = match.group(1)
    month = MONTH_ABBR[match.group(2).lower()]
    return f"{year}-{month:02d}"


def _rows_from_xlsx(content: bytes) -> List[List[str]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.worksheets[0]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    return rows


def _rows_from_csv(content: bytes) -> List[List[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def parse_workbook(content: bytes, filename: str) -> ImportResult:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
    elif filename.lower().endswith((".csv", ".tsv")):
        rows = _rows_from_csv(content)
    else:
        raise ValueError("Unsupported file type - upload an .xlsx or .csv export of the budget sheet")

    result = ImportResult()
    current_block: Optional[str] = None  # "goal" / "actual", once a GOALS/ACTUAL header row is seen

    for row in rows:
        tokens = [c.strip() for c in row if c is not None and str(c).strip() != ""]
        i = 0
        while i < len(tokens):
            token = tokens[i]
            upper = token.upper()

            if upper in ("GOALS", "ACTUAL"):
                current_block = "goal" if upper == "GOALS" else "actual"
                i += 1
                continue

            if upper in GOAL_TAGS and current_block is not None and i + 1 < len(tokens):
                pct = _parse_percent(tokens[i + 1])
                consumed = 2
                if i + 2 < len(tokens) and parse_european_amount(tokens[i + 2]) is not None:
                    consumed = 3  # trailing absolute amount column, currently unused but consumed so it isn't misread as a new row start
                if pct is not None:
                    goal = result.goals.setdefault(upper.capitalize(), GoalPct())
                    if current_block == "goal":
                        goal.target_pct = pct
                    else:
                        goal.actual_pct = pct
                i += consumed
                continue

            if upper in ("TOTAL INCOME", "TOTAL EXPENSES", "RAMAS") and i + 1 < len(tokens):
                i += 2  # computed from the rows themselves, not trusted from the sheet
                continue

            # name + tag, no amount (a handful of real rows have a blank amount cell)
            if i + 1 < len(tokens) and tokens[i + 1].capitalize() in CATEGORY_TAGS:
                result.warnings.append(f"\"{token}\" has no amount - skipped")
                i += 2
                continue

            # name + amount [+ tag]
            if i + 1 < len(tokens):
                amount = parse_european_amount(tokens[i + 1])
                if amount is not None:
                    if i + 2 < len(tokens) and tokens[i + 2].capitalize() in CATEGORY_TAGS:
                        result.categories.append(CategoryRow(name=token, amount=amount, tag=tokens[i + 2].capitalize()))
                        i += 3
                    else:
                        result.income.append(IncomeRow(name=token, amount=amount))
                        i += 2
                    continue

            # unparseable token (noise row, or a label we don't model) - skip just this one
            i += 1

    return result


def _parse_percent(token: str) -> Optional[float]:
    token = token.strip()
    if not token.endswith("%"):
        return None
    return parse_european_amount(token[:-1])
